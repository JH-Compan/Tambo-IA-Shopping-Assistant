"""
Script para generar los archivos Excel de prueba del proyecto Tambo.
Ejecutar una sola vez para crear los datos iniciales.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def estilo_encabezado(ws, columnas):
    for col, titulo in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="D4380D")  # rojo Tambo
        celda.alignment = Alignment(horizontal="center")

# ─────────────────────────────────────────────
# 1. productos.xlsx
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"
cols = ["id_producto", "nombre", "categoria", "precio", "descripcion"]
estilo_encabezado(ws, cols)

productos = [
    (1,  "Inca Kola 500ml",        "Bebidas",      3.50,  "Gaseosa nacional 500ml"),
    (2,  "Coca Cola 500ml",         "Bebidas",      3.50,  "Gaseosa internacional 500ml"),
    (3,  "Agua San Luis 625ml",     "Bebidas",      1.50,  "Agua mineral sin gas"),
    (4,  "Sporade Tropical 500ml",  "Bebidas",      2.50,  "Bebida rehidratante"),
    (5,  "Leche Gloria 1L",         "Lacteos",      5.00,  "Leche evaporada entera 1L"),
    (6,  "Yogurt Gloria Fresa",     "Lacteos",      4.00,  "Yogurt con fresa 1kg"),
    (7,  "Pan de Molde Bimbo",      "Panaderia",    6.50,  "Pan de molde blanco grande"),
    (8,  "Galletas Oreo",           "Snacks",       3.00,  "Galletas de chocolate con crema"),
    (9,  "Papas Lays Clasica",      "Snacks",       3.50,  "Papas fritas sabor original 70g"),
    (10, "Chocolate Sublime",       "Snacks",       2.00,  "Chocolate con mani"),
    (11, "Arroz Costeño 1kg",       "Abarrotes",    5.50,  "Arroz blanco extra 1kg"),
    (12, "Aceite Primor 1L",        "Abarrotes",    9.00,  "Aceite vegetal 1L"),
    (13, "Fideos Molitalia 500g",   "Abarrotes",    3.50,  "Fideos spaghetti 500g"),
    (14, "Detergente Ariel 500g",   "Limpieza",     8.50,  "Detergente en polvo 500g"),
    (15, "Jabon Dove 90g",          "Higiene",      4.50,  "Jabón de tocador hidratante"),
]
for fila in productos:
    ws.append(fila)

wb.save(os.path.join(BASE_DIR, "productos.xlsx"))
print("✅ productos.xlsx creado")

# ─────────────────────────────────────────────
# 2. stock.xlsx
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Stock"
cols = ["id_producto", "nombre", "stock_disponible"]
estilo_encabezado(ws, cols)

stock = [
    (1,  "Inca Kola 500ml",        50),
    (2,  "Coca Cola 500ml",         40),
    (3,  "Agua San Luis 625ml",     100),
    (4,  "Sporade Tropical 500ml",  30),
    (5,  "Leche Gloria 1L",         25),
    (6,  "Yogurt Gloria Fresa",     20),
    (7,  "Pan de Molde Bimbo",      15),
    (8,  "Galletas Oreo",           60),
    (9,  "Papas Lays Clasica",      0),   # agotado
    (10, "Chocolate Sublime",       45),
    (11, "Arroz Costeño 1kg",       35),
    (12, "Aceite Primor 1L",        10),
    (13, "Fideos Molitalia 500g",   55),
    (14, "Detergente Ariel 500g",   20),
    (15, "Jabon Dove 90g",          30),
]
for fila in stock:
    ws.append(fila)

wb.save(os.path.join(BASE_DIR, "stock.xlsx"))
print("✅ stock.xlsx creado")

# ─────────────────────────────────────────────
# 3. promociones.xlsx
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Promociones"
cols = ["id_promo", "nombre", "descripcion", "precio_promo", "fecha_inicio", "fecha_fin", "id_producto"]
estilo_encabezado(ws, cols)

hoy = date.today()
promos = [
    (1, "2x1 Inca Kola",       "Lleva 2 Inca Kola 500ml al precio de 1",    3.50, str(hoy), str(hoy + timedelta(days=7)),  1),
    (2, "Combo Snack",         "Lays + Oreo + Sublime por precio especial",  7.00, str(hoy), str(hoy + timedelta(days=5)),  8),
    (3, "Promo Desayuno",      "Leche Gloria + Pan de Molde juntos",         10.00,str(hoy), str(hoy + timedelta(days=3)),  5),
    (4, "Agua x6 unidades",    "Six pack de Agua San Luis 625ml",            8.00, str(hoy), str(hoy + timedelta(days=10)), 3),
    (5, "Promo Limpieza Hogar","Ariel 500g + Dove 90g juntos",               12.00,str(hoy), str(hoy + timedelta(days=4)), 14),
]
for fila in promos:
    ws.append(fila)

wb.save(os.path.join(BASE_DIR, "promociones.xlsx"))
print("✅ promociones.xlsx creado")

# ─────────────────────────────────────────────
# 4. historial_compras.xlsx
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Historial"
cols = ["id_compra", "telefono_cliente", "id_producto", "nombre_producto", "cantidad", "precio_unitario", "fecha_compra"]
estilo_encabezado(ws, cols)

historial = [
    (1,  "999111222", 1,  "Inca Kola 500ml",      2, 3.50, "2026-05-01"),
    (2,  "999111222", 5,  "Leche Gloria 1L",       1, 5.00, "2026-05-01"),
    (3,  "999111222", 7,  "Pan de Molde Bimbo",    1, 6.50, "2026-05-08"),
    (4,  "999111222", 1,  "Inca Kola 500ml",       2, 3.50, "2026-05-15"),
    (5,  "999111222", 8,  "Galletas Oreo",         2, 3.00, "2026-05-15"),
    (6,  "999333444", 11, "Arroz Costeño 1kg",     2, 5.50, "2026-05-10"),
    (7,  "999333444", 12, "Aceite Primor 1L",      1, 9.00, "2026-05-10"),
    (8,  "999333444", 13, "Fideos Molitalia 500g", 3, 3.50, "2026-05-20"),
    (9,  "999555666", 2,  "Coca Cola 500ml",       4, 3.50, "2026-05-18"),
    (10, "999555666", 10, "Chocolate Sublime",     2, 2.00, "2026-05-18"),
]
for fila in historial:
    ws.append(fila)

wb.save(os.path.join(BASE_DIR, "historial_compras.xlsx"))
print("✅ historial_compras.xlsx creado")

# ─────────────────────────────────────────────
# 5. interacciones.xlsx
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Interacciones"
cols = ["id_interaccion", "telefono_cliente", "mensaje_usuario", "respuesta_bot", "timestamp"]
estilo_encabezado(ws, cols)

wb.save(os.path.join(BASE_DIR, "interacciones.xlsx"))
print("✅ interacciones.xlsx creado (vacío, se llenará automáticamente)")
print("\n🎉 Todos los archivos Excel generados correctamente en /data/")
